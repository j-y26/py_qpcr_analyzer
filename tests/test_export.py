"""Tests for core/export.py — sheet naming, sample-name parity, CSV-zip."""

import io
import zipfile
from openpyxl import load_workbook

import pandas as pd
import pytest

from qpcr_analyzer.core.export import (
    _build_formatted_table,
    results_to_csv_zip_bytes,
    results_to_xlsx_bytes,
)


def _raw():
    return pd.DataFrame(
        {
            "Well": ["A1", "A2", "A3", "A4"],
            "Target": ["geneA", "geneB", "geneA", "geneB"],
            "Sample": ["s1", "s1", "s2", "s2"],
            "Group": ["ctrl", "ctrl", "treat", "treat"],
            "Cq": [20.0, 22.0, 21.0, 23.0],
        }
    )


def _dct_results():
    return {
        "geneA": pd.DataFrame(
            {
                "Target": ["geneB", "geneB"],
                "Group": ["ctrl", "treat"],
                "Sample": ["s1", "s2"],
                "Mean_Cq": [22.0, 23.0],
                "Mean_Cq_geneA": [20.0, 21.0],
                "dCt": [2.0, 2.0],
                "Expr_vs_HK": [0.25, 0.25],
                "Reference_Gene": ["geneA", "geneA"],
            }
        )
    }


def _ddct_results():
    return {
        "geneA": pd.DataFrame(
            {
                "Target": ["geneB", "geneB"],
                "Group": ["ctrl", "treat"],
                "Sample": ["s1", "s2"],
                "Batch": ["batch_1", "batch_1"],
                "Mean_Cq": [22.0, 23.0],
                "Mean_Cq_geneA": [20.0, 21.0],
                "dCt": [2.0, 2.0],
                "Ref_dCt": [2.0, 2.0],
                "ddCt": [0.0, 0.0],
                "Relative_Expr": [1.0, 1.0],
                "Is_Reference_Group": [True, False],
                "Reference_Gene": ["geneA", "geneA"],
                "Reference_Group": ["ctrl", "ctrl"],
            }
        )
    }


def test_xlsx_uses_formatted_prefix_not_prism():
    data = results_to_xlsx_bytes(_raw(), _dct_results(), _ddct_results())
    wb = load_workbook(io.BytesIO(data), read_only=True)
    names = set(wb.sheetnames)
    # New naming
    assert any(n.startswith("formatted_dCt_") for n in names)
    assert any(n.startswith("formatted_ddCt_") for n in names)
    # Old naming retired
    assert not any(n.startswith("prism_") for n in names)


def test_formatted_sheet_includes_sample_columns():
    df = _dct_results()["geneA"]
    formatted = _build_formatted_table(df, value_col="dCt")
    # One Sample column per group, plus Target
    assert "Target" in formatted.columns
    assert "Sample (ctrl)" in formatted.columns
    assert "Sample (treat)" in formatted.columns
    assert "Relative Expression (ctrl)" in formatted.columns
    assert "Relative Expression (treat)" in formatted.columns

    # The data row(s) contain the sample names alongside values
    data_rows = formatted[formatted["Target"] == ""]
    sample_cells = (
        data_rows["Sample (ctrl)"].tolist() + data_rows["Sample (treat)"].tolist()
    )
    assert "s1" in sample_cells
    assert "s2" in sample_cells


def test_formatted_parity_check_passes_on_consistent_data():
    # If parity were broken, _build_formatted_table itself would raise.
    out = _build_formatted_table(_dct_results()["geneA"], value_col="dCt")
    assert not out.empty


def test_formatted_parity_check_raises_on_synthetic_mismatch(monkeypatch):
    """Tampering with the verification data must surface as ValueError."""
    from qpcr_analyzer.core import export as exp

    df = _dct_results()["geneA"].copy()

    real_verify = exp._verify_formatted_against_source

    def bad_verify(written, source, value_col):
        # Swap the first written tuple's value to provoke a mismatch.
        if written:
            t, g, s, _v = written[0]
            written[0] = (t, g, s, 9999.99)
        return real_verify(written, source, value_col)

    monkeypatch.setattr(exp, "_verify_formatted_against_source", bad_verify)
    with pytest.raises(ValueError, match="parity check failed"):
        exp._build_formatted_table(df, value_col="dCt")


def test_csv_zip_contains_all_sheets():
    data = results_to_csv_zip_bytes(_raw(), _dct_results(), _ddct_results())
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = set(zf.namelist())
    assert "raw_data.csv" in names
    assert any(n.startswith("dCt_") and n.endswith(".csv") for n in names)
    assert any(n.startswith("ddCt_") and n.endswith(".csv") for n in names)
    assert any(n.startswith("formatted_dCt_") and n.endswith(".csv") for n in names)
    assert any(n.startswith("formatted_ddCt_") and n.endswith(".csv") for n in names)


def test_csv_zip_csv_content_matches_xlsx_sheets():
    """Round-trip a CSV from the zip and compare against the source frame."""
    data = results_to_csv_zip_bytes(_raw(), _dct_results(), _ddct_results())
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        with zf.open("dCt_geneA.csv") as f:
            csv_df = pd.read_csv(f)
    src = _dct_results()["geneA"]
    assert list(csv_df["Sample"]) == list(src["Sample"])
    assert list(csv_df["dCt"]) == list(src["dCt"])
